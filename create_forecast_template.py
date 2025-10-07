#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генератор Excel-шаблона для расчёта метрик прогнозирования
Forecast_Metrics_Template.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.comments import Comment
from datetime import datetime, timedelta
import random

# Константы цветов
COLOR_INPUT = "E2F0D9"  # Светло-зелёный для вводимых полей
COLOR_CALC = "F2F2F2"   # Серый для вычисляемых полей
COLOR_HEADER = "404040"  # Тёмно-серый для заголовков
COLOR_WHITE = "FFFFFF"
COLOR_YELLOW = "FFEB9C"  # Предупреждение
COLOR_ORANGE = "FFC000"  # Выбросы
COLOR_RED = "FF0000"     # Критично
COLOR_AMBER = "FFA500"   # Янтарный
COLOR_GREEN = "00B050"   # Хорошо
COLOR_BLUE = "4472C4"    # Недопрогноз
COLOR_PURPLE = "7030A0"  # Перепрогноз

def create_workbook():
    """Создание основной структуры книги"""
    wb = Workbook()
    
    # Удаляем стандартный лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Создаём листы в нужном порядке
    ws_instructions = wb.create_sheet("Инструкции", 0)
    ws_input = wb.create_sheet("Ввод_данных", 1)
    ws_metrics = wb.create_sheet("Метрики", 2)
    ws_charts = wb.create_sheet("Графики", 3)
    
    return wb, ws_instructions, ws_input, ws_metrics, ws_charts

def style_header(cell):
    """Стиль для заголовка"""
    cell.font = Font(name='Calibri', size=11, bold=True, color=COLOR_WHITE)
    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def style_input(cell):
    """Стиль для ячейки ввода"""
    cell.font = Font(name='Calibri', size=11)
    cell.fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_calc(cell):
    """Стиль для вычисляемой ячейки"""
    cell.font = Font(name='Calibri', size=11)
    cell.fill = PatternFill(start_color=COLOR_CALC, end_color=COLOR_CALC, fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')

def create_instructions_sheet(ws):
    """Создание листа Инструкции"""
    ws.column_dimensions['A'].width = 100
    
    # Заголовок
    ws['A1'] = "ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ ШАБЛОНА FORECAST METRICS"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=COLOR_HEADER)
    ws.merge_cells('A1:A1')
    
    row = 3
    
    # Секция 1: Назначение
    ws[f'A{row}'] = "1. НАЗНАЧЕНИЕ ФАЙЛА"
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color=COLOR_HEADER)
    row += 1
    
    ws[f'A{row}'] = ("Этот шаблон предназначен для расчёта и анализа точности прогнозов. "
                     "Он вычисляет ключевые метрики: MAPE (средняя процентная ошибка), "
                     "WAPE (взвешенная процентная ошибка), MAE (средняя абсолютная ошибка), "
                     "RMSE (среднеквадратичная ошибка) и Bias (смещение прогноза).")
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    row += 2
    
    # Секция 2: Как работать
    ws[f'A{row}'] = "2. КАК РАБОТАТЬ С ШАБЛОНОМ"
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color=COLOR_HEADER)
    row += 1
    
    instructions = [
        "Шаг 1: Перейдите на лист 'Ввод_данных'",
        "Шаг 2: Заполните колонки 'SKU', 'Период', 'Факт_y' и 'Прогноз_f' (светло-зелёные ячейки)",
        "Шаг 3: Проверьте автоматическую подсветку ошибок и нулевых значений",
        "Шаг 4: Перейдите на лист 'Метрики' для просмотра рассчитанных показателей",
        "Шаг 5: Изучите диагностику и рекомендации",
        "Шаг 6: Посмотрите визуализацию на листе 'Графики'"
    ]
    
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    row += 1
    
    # Секция 3: Что такое метрики
    ws[f'A{row}'] = "3. ЧТО ТАКОЕ МЕТРИКИ"
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color=COLOR_HEADER)
    row += 1
    
    metrics_info = [
        "MAPE (Mean Absolute Percentage Error) — средняя относительная ошибка в процентах. "
        "Показывает, насколько в среднем прогноз отклоняется от факта. "
        "Применяется для стабильных рядов без нулевых значений.",
        
        "WAPE (Weighted Absolute Percentage Error) — взвешенная ошибка по объёму. "
        "Учитывает важность больших объёмов продаж. Идеально для портфелей SKU и отчётности.",
        
        "MAE (Mean Absolute Error) — средняя абсолютная ошибка в штуках. "
        "Простая и понятная метрика, показывает среднее отклонение в единицах товара.",
        
        "RMSE (Root Mean Square Error) — среднеквадратичная ошибка. "
        "Усиливает влияние больших промахов. Если RMSE >> MAE, значит есть выбросы.",
        
        "Bias — систематическое смещение прогноза. "
        "Положительное значение = недопрогноз (риск дефицитов), "
        "отрицательное = перепрогноз (риск излишков)."
    ]
    
    for info in metrics_info:
        ws[f'A{row}'] = info
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 40
        row += 1
    row += 1
    
    # Секция 4: Ограничения
    ws[f'A{row}'] = "4. ОГРАНИЧЕНИЯ И ВАЖНЫЕ ЗАМЕЧАНИЯ"
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color=COLOR_HEADER)
    row += 1
    
    limitations = [
        "⚠ MAPE не рассчитывается при нулевых фактических значениях (деление на ноль)",
        "⚠ Если доля нулей >30%, не используйте MAPE — применяйте WAPE или MAE",
        "⚠ RMSE чувствителен к выбросам — сравнивайте с MAE для диагностики",
        "⚠ Bias показывает направление, но не величину разброса — смотрите вместе с другими метриками",
        "⚠ При пустых или неполных данных метрики вернут #Н/Д (NA)"
    ]
    
    for limitation in limitations:
        ws[f'A{row}'] = limitation
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    row += 1
    
    # Секция 5: Легенда цветов
    ws[f'A{row}'] = "5. ЛЕГЕНДА ЦВЕТОВ"
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color=COLOR_HEADER)
    row += 1
    
    # Примеры цветов
    color_legend = [
        ("Светло-зелёный", COLOR_INPUT, "Поля для ввода данных"),
        ("Серый", COLOR_CALC, "Вычисляемые поля (защищены)"),
        ("Жёлтый", COLOR_YELLOW, "Предупреждение (y=0, MAPE не считается)"),
        ("Оранжевый", COLOR_ORANGE, "Выбросы (топ-10% ошибок)"),
        ("Зелёный", COLOR_GREEN, "WAPE <10% (хорошо)"),
        ("Янтарный", COLOR_AMBER, "WAPE 10-15% (приемлемо)"),
        ("Красный", COLOR_RED, "WAPE >15% (требует внимания)")
    ]
    
    for legend_name, color, description in color_legend:
        ws[f'A{row}'] = f"{legend_name}: {description}"
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    row += 1
    
    # Секция 6: Как читать результаты
    ws[f'A{row}'] = "6. КАК ЧИТАТЬ РЕЗУЛЬТАТЫ"
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color=COLOR_HEADER)
    row += 1
    
    reading_rules = [
        "✓ MAPE — используйте для стабильных рядов без нулей и выбросов",
        "✓ WAPE — основная метрика для портфелей и отчётности руководству",
        "✓ RMSE vs MAE — если RMSE значительно больше MAE, значит присутствуют выбросы",
        "✓ Bias — всегда проверяйте знак: + (недопрогноз, дефициты), - (перепрогноз, излишки)",
        "✓ Если доля нулей >30%, не используйте MAPE"
    ]
    
    for rule in reading_rules:
        ws[f'A{row}'] = rule
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Заморозить верхнюю строку
    ws.freeze_panes = 'A2'

def create_input_sheet(ws):
    """Создание листа Ввод_данных"""
    
    # Заголовки
    headers = [
        ("A1", "SKU", "Артикул или название товара"),
        ("B1", "Период", "Дата или номер периода"),
        ("C1", "Факт_y (шт)", "Фактические продажи"),
        ("D1", "Прогноз_f (шт)", "Прогнозные продажи"),
        ("E1", "Ошибка_e", "Разница: Факт - Прогноз"),
        ("F1", "|e|", "Абсолютная ошибка"),
        ("G1", "%AbsErr", "Процент абсолютной ошибки"),
        ("H1", "Примечание", "Дополнительная информация")
    ]
    
    for cell_ref, header_text, comment_text in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        style_header(cell)
        # Добавляем комментарий-подсказку
        cell.comment = Comment(comment_text, "System")
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 30
    
    # Формулы для строк 2-1000
    for row in range(2, 1001):
        # Стили для ввода (C, D, H)
        style_input(ws[f'C{row}'])
        style_input(ws[f'D{row}'])
        style_input(ws[f'H{row}'])
        
        # Стили для вычисляемых (E, F, G)
        style_calc(ws[f'E{row}'])
        style_calc(ws[f'F{row}'])
        style_calc(ws[f'G{row}'])
        
        # Формулы (используем ; для RU locale)
        ws[f'E{row}'] = f'=C{row}-D{row}'
        ws[f'F{row}'] = f'=ABS(E{row})'
        ws[f'G{row}'] = f'=IF(C{row}>0;F{row}/C{row};NA())'
        
        # Форматы чисел
        ws[f'C{row}'].number_format = '0'
        ws[f'D{row}'].number_format = '0'
        ws[f'E{row}'].number_format = '0'
        ws[f'F{row}'].number_format = '0'
        ws[f'G{row}'].number_format = '0.00%'
    
    # Валидация данных для C и D (только положительные числа)
    dv_positive = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0)
    dv_positive.error = 'Значение должно быть >= 0'
    dv_positive.errorTitle = 'Ошибка ввода'
    dv_positive.add('C2:C1000')
    dv_positive.add('D2:D1000')
    ws.add_data_validation(dv_positive)
    
    # Условное форматирование: жёлтая заливка при y=0
    yellow_fill = PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type='solid')
    rule_zero = FormulaRule(formula=['$C2=0'], fill=yellow_fill)
    ws.conditional_formatting.add('C2:C1000', rule_zero)
    
    # Условное форматирование: оранжевая заливка для топ-10% ошибок
    orange_fill = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type='solid')
    # Формула для топ-10%: если |e| >= 90-го перцентиля
    rule_outlier = FormulaRule(
        formula=[f'$F2>=PERCENTILE($F$2:$F$1000;0.9)'],
        fill=orange_fill
    )
    ws.conditional_formatting.add('F2:F1000', rule_outlier)
    
    # Заморозить верхнюю строку
    ws.freeze_panes = 'A2'

def create_metrics_sheet(ws):
    """Создание листа Метрики"""
    
    # Заголовок
    ws['A1'] = "РАССЧИТАННЫЕ МЕТРИКИ ПРОГНОЗА"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=COLOR_HEADER)
    ws.merge_cells('A1:E1')
    
    # Заголовки таблицы метрик
    headers = ["Метрика", "Значение", "Единицы", "Интерпретация", "Примечание"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        style_header(cell)
    
    # Данные метрик
    metrics_data = [
        ("MAPE", "=IFERROR(AVERAGE(IF(ISNUMBER(Ввод_данных!G2:G1000);Ввод_данных!G2:G1000));NA())", 
         "%", "Средняя относительная ошибка", "Не рассчитывается, где y=0"),
        
        ("WAPE", "=IF(SUM(Ввод_данных!C2:C1000)>0;SUM(Ввод_данных!F2:F1000)/SUM(Ввод_данных!C2:C1000);NA())", 
         "%", "Взвешенная ошибка по объёму", "Основная метрика для портфелей"),
        
        ("MAE", "=IFERROR(AVERAGE(IF(Ввод_данных!F2:F1000<>\"\";Ввод_данных!F2:F1000));NA())", 
         "шт", "Средняя абсолютная ошибка", "Простая и понятная метрика"),
        
        ("RMSE", "=IFERROR(SQRT(AVERAGE((Ввод_данных!E2:E1000)^2));NA())", 
         "шт", "Усиливает влияние больших промахов", "Сравните с MAE"),
        
        ("Bias", "=IFERROR(AVERAGE(IF(Ввод_данных!E2:E1000<>\"\";Ввод_данных!E2:E1000));NA())", 
         "шт", "Знак = направление смещения", "+ недопрогноз, - перепрогноз")
    ]
    
    for row, (metric, formula, unit, interp, note) in enumerate(metrics_data, start=4):
        ws.cell(row=row, column=1).value = metric
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        ws.cell(row=row, column=2).value = formula
        ws.cell(row=row, column=2).number_format = '0.00%' if unit == '%' else '0.00'
        style_calc(ws.cell(row=row, column=2))
        
        ws.cell(row=row, column=3).value = unit
        ws.cell(row=row, column=4).value = interp
        ws.cell(row=row, column=5).value = note
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
    
    # Именованные ячейки для удобства
    wb = ws.parent
    wb.create_named_range('MAPE_Value', ws, '$B$4')
    wb.create_named_range('WAPE_Value', ws, '$B$5')
    wb.create_named_range('MAE_Value', ws, '$B$6')
    wb.create_named_range('RMSE_Value', ws, '$B$7')
    wb.create_named_range('Bias_Value', ws, '$B$8')
    
    # Секция диагностики
    row = 10
    ws[f'A{row}'] = "ДИАГНОСТИКА И РЕКОМЕНДАЦИИ"
    ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color=COLOR_HEADER)
    ws.merge_cells(f'A{row}:E{row}')
    
    # Знак Bias
    row += 2
    ws[f'A{row}'] = "Знак Bias:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = '=IF(B8>0;"Недопрогноз (дефициты)";IF(B8<0;"Перепрогноз (излишки)";"Смещения нет"))'
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    style_calc(ws[f'B{row}'])
    ws.merge_cells(f'B{row}:E{row}')
    
    # Сравнение RMSE vs MAE
    row += 1
    ws[f'A{row}'] = "RMSE vs MAE:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = '=IF(B7>B6;"Есть выбросы/крупные промахи";"Ошибки равномерны")'
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    style_calc(ws[f'B{row}'])
    ws.merge_cells(f'B{row}:E{row}')
    
    # Рекомендации
    row += 2
    ws[f'A{row}'] = "РЕКОМЕНДАЦИИ:"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    recommendations = [
        ("Высокий WAPE (>15%):", '=IF(B5>0.15;"⚠ Проверьте сезонность, акции и качество данных";"✓ WAPE в норме")'),
        ("Разница RMSE и MAE:", '=IF(B7>B6*1.2;"⚠ Сгладьте выбросы или разнесите всплески";"✓ Выбросов нет")'),
        ("Значимый |Bias|:", '=IF(ABS(B8)>B6*0.3;"⚠ Сместите модель/перепараметризуйте";"✓ Bias незначителен")')
    ]
    
    for rec_label, rec_formula in recommendations:
        row += 1
        ws[f'A{row}'] = rec_label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = rec_formula
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        style_calc(ws[f'B{row}'])
        ws.merge_cells(f'B{row}:E{row}')
    
    # Условное форматирование для WAPE
    green_fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type='solid')
    amber_fill = PatternFill(start_color=COLOR_AMBER, end_color=COLOR_AMBER, fill_type='solid')
    red_fill = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type='solid')
    
    ws.conditional_formatting.add('B5', CellIsRule(operator='lessThan', formula=['0.10'], fill=green_fill))
    ws.conditional_formatting.add('B5', CellIsRule(operator='between', formula=['0.10', '0.15'], fill=amber_fill))
    ws.conditional_formatting.add('B5', CellIsRule(operator='greaterThan', formula=['0.15'], fill=red_fill))
    
    # Условное форматирование для Bias (синий/фиолетовый)
    blue_fill = PatternFill(start_color=COLOR_BLUE, end_color=COLOR_BLUE, fill_type='solid')
    purple_fill = PatternFill(start_color=COLOR_PURPLE, end_color=COLOR_PURPLE, fill_type='solid')
    
    ws.conditional_formatting.add('B8', CellIsRule(operator='greaterThan', formula=['0'], fill=blue_fill))
    ws.conditional_formatting.add('B8', CellIsRule(operator='lessThan', formula=['0'], fill=purple_fill))
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 35
    
    # Заморозить верхние строки
    ws.freeze_panes = 'A4'

def create_charts_sheet(ws, ws_input):
    """Создание листа Графики"""
    
    # Заголовок
    ws['A1'] = "ГРАФИКИ И ВИЗУАЛИЗАЦИЯ"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=COLOR_HEADER)
    ws.merge_cells('A1:H1')
    
    # График 1: Линейный график Факт vs Прогноз
    chart1 = LineChart()
    chart1.title = "Факт vs Прогноз"
    chart1.style = 13
    chart1.y_axis.title = 'Объём (шт)'
    chart1.x_axis.title = 'Период'
    chart1.width = 20
    chart1.height = 12
    
    # Данные для графика (первые 50 строк)
    data1 = Reference(ws_input, min_col=3, min_row=1, max_row=50, max_col=4)
    cats1 = Reference(ws_input, min_col=2, min_row=2, max_row=50)
    
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    
    ws.add_chart(chart1, "A3")
    
    # График 2: Столбчатый график абсолютных ошибок
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Абсолютная ошибка по периодам"
    chart2.style = 10
    chart2.y_axis.title = 'Абсолютная ошибка (шт)'
    chart2.x_axis.title = 'Период'
    chart2.width = 20
    chart2.height = 12
    
    data2 = Reference(ws_input, min_col=6, min_row=1, max_row=50)
    cats2 = Reference(ws_input, min_col=2, min_row=2, max_row=50)
    
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    
    ws.add_chart(chart2, "A20")
    
    # Примечание
    ws['A37'] = "Примечание: Графики автоматически обновляются при изменении данных на листе 'Ввод_данных'"
    ws['A37'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A37:H37')

def add_test_data(ws):
    """Добавление тестовых данных (12 периодов)"""
    
    # SKU
    sku = "TEST-SKU-001"
    
    # Стартовая дата
    start_date = datetime(2024, 1, 1)
    
    # Генерация тестовых данных с WAPE около 8-10%
    test_data = []
    base_demand = 100
    
    for i in range(12):
        period = start_date + timedelta(days=30*i)
        # Сезонность
        seasonal_factor = 1 + 0.3 * (i % 4 - 1.5) / 1.5
        actual = int(base_demand * seasonal_factor + random.uniform(-10, 10))
        # Прогноз с небольшой ошибкой
        forecast = int(actual + random.uniform(-12, 8))
        
        test_data.append({
            'sku': sku,
            'period': period,
            'actual': max(1, actual),
            'forecast': max(1, forecast)
        })
    
    # Записываем данные
    for i, data in enumerate(test_data, start=2):
        ws[f'A{i}'] = data['sku']
        ws[f'B{i}'] = data['period']
        ws[f'B{i}'].number_format = 'DD.MM.YYYY'
        ws[f'C{i}'].value = data['actual']
        ws[f'D{i}'].value = data['forecast']

def protect_sheets(wb):
    """Защита листов"""
    
    # Защищаем листы Метрики и Графики полностью
    wb['Метрики'].protection.sheet = True
    wb['Метрики'].protection.password = 'forecast2024'
    
    wb['Графики'].protection.sheet = True
    wb['Графики'].protection.password = 'forecast2024'
    
    # Защищаем лист Ввод_данных, но разрешаем редактирование C, D, H
    ws_input = wb['Ввод_данных']
    ws_input.protection.sheet = True
    ws_input.protection.password = 'forecast2024'
    
    # Разблокируем ячейки для ввода (A, B, C, D, H)
    for row in range(2, 1001):
        ws_input[f'A{row}'].protection = openpyxl.styles.Protection(locked=False)
        ws_input[f'B{row}'].protection = openpyxl.styles.Protection(locked=False)
        ws_input[f'C{row}'].protection = openpyxl.styles.Protection(locked=False)
        ws_input[f'D{row}'].protection = openpyxl.styles.Protection(locked=False)
        ws_input[f'H{row}'].protection = openpyxl.styles.Protection(locked=False)

def main():
    """Главная функция"""
    print("Создание Excel-шаблона Forecast_Metrics_Template.xlsx...")
    
    # Создаём книгу и листы
    wb, ws_instructions, ws_input, ws_metrics, ws_charts = create_workbook()
    
    print("Создание листа 'Инструкции'...")
    create_instructions_sheet(ws_instructions)
    
    print("Создание листа 'Ввод_данных'...")
    create_input_sheet(ws_input)
    
    print("Добавление тестовых данных...")
    add_test_data(ws_input)
    
    print("Создание листа 'Метрики'...")
    create_metrics_sheet(ws_metrics)
    
    print("Создание листа 'Графики'...")
    create_charts_sheet(ws_charts, ws_input)
    
    print("Настройка защиты листов...")
    protect_sheets(wb)
    
    # Сохраняем файл
    filename = "Forecast_Metrics_Template.xlsx"
    wb.save(filename)
    print(f"✓ Файл '{filename}' успешно создан!")
    
    return filename

if __name__ == "__main__":
    import openpyxl.styles
    main()
