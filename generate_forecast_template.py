#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генератор Excel-шаблона для расчёта метрик прогнозирования
Forecast_Metrics_Template.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import date, timedelta
import openpyxl

def create_forecast_template():
    """Создание Excel-шаблона согласно ТЗ"""
    
    # Создаем книгу
    wb = Workbook()
    
    # Удаляем стандартный лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Цвета
    COLOR_INPUT = "E2F0D9"  # светло-зелёный
    COLOR_CALC = "F2F2F2"   # серый
    COLOR_HEADER = "404040"  # тёмно-серый
    COLOR_WARNING = "FFFF00"  # жёлтый
    COLOR_ERROR = "FFA500"    # оранжевый
    COLOR_GREEN = "92D050"    # зелёный
    COLOR_AMBER = "FFC000"    # янтарный
    COLOR_RED = "FF0000"      # красный
    COLOR_BLUE = "0070C0"     # синий
    COLOR_PURPLE = "7030A0"   # фиолетовый
    
    # Стили
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    normal_font = Font(name='Calibri', size=11)
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    input_fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type='solid')
    calc_fill = PatternFill(start_color=COLOR_CALC, end_color=COLOR_CALC, fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============================================
    # Лист 1: Инструкции
    # ============================================
    ws_instr = wb.create_sheet("Инструкции", 0)
    
    instructions = [
        ["ШАБЛОН ДЛЯ РАСЧЁТА МЕТРИК ПРОГНОЗИРОВАНИЯ", ""],
        ["", ""],
        ["1. НАЗНАЧЕНИЕ ФАЙЛА", ""],
        ["", "Этот шаблон предназначен для расчёта и интерпретации ключевых метрик точности прогнозирования:"],
        ["", "• MAPE (Mean Absolute Percentage Error) — средняя абсолютная процентная ошибка"],
        ["", "• WAPE (Weighted Absolute Percentage Error) — взвешенная абсолютная процентная ошибка"],
        ["", "• MAE (Mean Absolute Error) — средняя абсолютная ошибка"],
        ["", "• RMSE (Root Mean Square Error) — среднеквадратичная ошибка"],
        ["", "• Bias — систематическое смещение прогноза"],
        ["", ""],
        ["2. КАК РАБОТАТЬ С ШАБЛОНОМ", ""],
        ["", "Шаг 1: Перейдите на лист «Ввод_данных» и вставьте ваши данные в колонки C (Факт) и D (Прогноз)"],
        ["", "Шаг 2: Проверьте подсветку ошибок (жёлтые ячейки = нули, оранжевые = выбросы)"],
        ["", "Шаг 3: Перейдите на лист «Метрики» для просмотра рассчитанных показателей"],
        ["", "Шаг 4: Изучите визуализацию на листе «Графики»"],
        ["", ""],
        ["3. ЧТО ТАКОЕ МЕТРИКИ И КОГДА ИХ ПРИМЕНЯТЬ", ""],
        ["", "MAPE — показывает среднюю относительную ошибку в %. Используйте для стабильных рядов без нулей."],
        ["", "WAPE — взвешенная ошибка по объёму. Лучше подходит для портфелей и отчётности."],
        ["", "MAE — средняя абсолютная ошибка в штуках. Простая интерпретация, устойчива к выбросам."],
        ["", "RMSE — усиливает влияние больших промахов. Сравните с MAE для выявления выбросов."],
        ["", "Bias — показывает направление смещения: + недопрогноз (дефициты), - перепрогноз (излишки)."],
        ["", ""],
        ["4. ОГРАНИЧЕНИЯ И ВАЖНЫЕ МОМЕНТЫ", ""],
        ["", "• MAPE не рассчитывается при нулевых фактических значениях (y=0)"],
        ["", "• RMSE чувствителен к выбросам — большие промахи влияют сильнее"],
        ["", "• Bias всегда нужно проверять вместе с MAE/WAPE для полной картины"],
        ["", "• Если доля нулей >30%, не используйте MAPE — используйте WAPE"],
        ["", ""],
        ["5. ЛЕГЕНДА ЦВЕТОВ", ""],
        ["", "🟢 Светло-зелёный — поля для ввода данных (можно редактировать)"],
        ["", "⚪ Серый — вычисляемые поля (защищены, формулы)"],
        ["", "⚫ Тёмно-серый — заголовки таблиц"],
        ["", "🟡 Жёлтый — предупреждение (нулевые значения факта)"],
        ["", "🟠 Оранжевый — выбросы (топ-10% ошибок)"],
        ["", ""],
        ["6. КАК ЧИТАТЬ РЕЗУЛЬТАТЫ", ""],
        ["", "• WAPE < 10% — отличный прогноз; 10-15% — хороший; >15% — требуется улучшение"],
        ["", "• RMSE > MAE — присутствуют выбросы или крупные промахи"],
        ["", "• Bias > 0 — систематический недопрогноз (риск дефицитов)"],
        ["", "• Bias < 0 — систематический перепрогноз (риск излишков)"],
        ["", "• Высокий WAPE → проверьте сезонность, акции, качество исходных данных"],
    ]
    
    for i, row in enumerate(instructions, 1):
        ws_instr.cell(row=i, column=1, value=row[0])
        ws_instr.cell(row=i, column=2, value=row[1])
        
        if row[0].startswith(("1.", "2.", "3.", "4.", "5.", "6.")):
            ws_instr.cell(row=i, column=1).font = Font(name='Calibri', size=12, bold=True)
            ws_instr.cell(row=i, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type='solid')
        elif "ШАБЛОН" in row[0]:
            ws_instr.cell(row=i, column=1).font = Font(name='Calibri', size=14, bold=True)
            ws_instr.cell(row=i, column=1).alignment = Alignment(horizontal='center')
            ws_instr.merge_cells(f'A{i}:B{i}')
    
    ws_instr.column_dimensions['A'].width = 40
    ws_instr.column_dimensions['B'].width = 80
    
    # ============================================
    # Лист 2: Ввод_данных
    # ============================================
    ws_data = wb.create_sheet("Ввод_данных", 1)
    
    # Заголовки
    headers = [
        ("SKU", "Артикул или код товара"),
        ("Период", "Дата или номер периода"),
        ("Факт_y", "Фактические продажи (шт)"),
        ("Прогноз_f", "Прогнозные значения (шт)"),
        ("Ошибка_e", "Разница: Факт - Прогноз"),
        ("Абс_Ошибка", "Модуль ошибки |e|"),
        ("Проц_Ошибка", "Абсолютная процентная ошибка |e|/y"),
        ("Примечание", "Дополнительная информация")
    ]
    
    for col, (header, comment_text) in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
        # Добавляем примечание-подсказку
        cell.comment = Comment(comment_text, "Шаблон")
    
    # Форматирование колонок
    ws_data.column_dimensions['A'].width = 15
    ws_data.column_dimensions['B'].width = 12
    ws_data.column_dimensions['C'].width = 12
    ws_data.column_dimensions['D'].width = 12
    ws_data.column_dimensions['E'].width = 12
    ws_data.column_dimensions['F'].width = 12
    ws_data.column_dimensions['G'].width = 14
    ws_data.column_dimensions['H'].width = 25
    
    # Заполняем формулы для строк 2-1000
    for row in range(2, 1001):
        # A и B - пустые для ввода
        ws_data.cell(row=row, column=1).fill = input_fill  # SKU
        ws_data.cell(row=row, column=2).fill = input_fill  # Период
        
        # C и D - поля ввода
        ws_data.cell(row=row, column=3).fill = input_fill  # Факт
        ws_data.cell(row=row, column=4).fill = input_fill  # Прогноз
        
        # E - Ошибка = C - D
        ws_data.cell(row=row, column=5, value=f'=C{row}-D{row}')
        ws_data.cell(row=row, column=5).fill = calc_fill
        
        # F - |Ошибка|
        ws_data.cell(row=row, column=6, value=f'=ABS(E{row})')
        ws_data.cell(row=row, column=6).fill = calc_fill
        
        # G - % Ошибка с защитой от деления на ноль
        ws_data.cell(row=row, column=7, value=f'=IF(C{row}>0;F{row}/C{row};NA())')
        ws_data.cell(row=row, column=7).fill = calc_fill
        ws_data.cell(row=row, column=7).number_format = '0.00%'
        
        # H - Примечание
        ws_data.cell(row=row, column=8).fill = input_fill
    
    # Валидация данных: C и D >= 0
    dv_positive = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0, 
                                  allow_blank=True, showErrorMessage=True,
                                  error='Значение должно быть >= 0', errorTitle='Ошибка ввода')
    dv_positive.add('C2:C1000')
    dv_positive.add('D2:D1000')
    ws_data.add_data_validation(dv_positive)
    
    # Условное форматирование: подсветка нулей в колонке C (жёлтый)
    ws_data.conditional_formatting.add('C2:C1000',
        FormulaRule(formula=['C2=0'], fill=PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type='solid')))
    
    # Условное форматирование: выбросы (топ 10%) в колонке F (оранжевый)
    ws_data.conditional_formatting.add('F2:F1000',
        FormulaRule(formula=[f'F2>=PERCENTILE($F$2:$F$1000,0.9)'], 
                    fill=PatternFill(start_color=COLOR_ERROR, end_color=COLOR_ERROR, fill_type='solid')))
    
    # Заморозить верхнюю строку
    ws_data.freeze_panes = 'A2'
    
    # ============================================
    # Именованные диапазоны
    # ============================================
    from openpyxl.workbook.defined_name import DefinedName
    
    # Создаём именованные диапазоны с использованием современного API
    wb.defined_names['rngY'] = DefinedName('rngY', attr_text=f"Ввод_данных!$C$2:$C$1000")
    wb.defined_names['rngF'] = DefinedName('rngF', attr_text=f"Ввод_данных!$D$2:$D$1000")
    wb.defined_names['rngE'] = DefinedName('rngE', attr_text=f"Ввод_данных!$E$2:$E$1000")
    wb.defined_names['rngAbsE'] = DefinedName('rngAbsE', attr_text=f"Ввод_данных!$F$2:$F$1000")
    wb.defined_names['rngPctAbsE'] = DefinedName('rngPctAbsE', attr_text=f"Ввод_данных!$G$2:$G$1000")
    wb.defined_names['rngSKU'] = DefinedName('rngSKU', attr_text=f"Ввод_данных!$A$2:$A$1000")
    
    # ============================================
    # Лист 3: Метрики
    # ============================================
    ws_metrics = wb.create_sheet("Метрики", 2)
    
    # Заголовок
    ws_metrics.merge_cells('A1:E1')
    ws_metrics['A1'] = 'РАСЧЁТ МЕТРИК ПРОГНОЗИРОВАНИЯ'
    ws_metrics['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws_metrics['A1'].alignment = center_align
    
    # Таблица метрик
    metric_headers = ["Метрика", "Значение", "Единицы", "Интерпретация", "Примечание"]
    for col, header in enumerate(metric_headers, 1):
        cell = ws_metrics.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Данные метрик
    metrics_data = [
        ("MAPE", "=IFERROR(AVERAGE(IF(ISNUMBER(rngPctAbsE);rngPctAbsE));NA())", "%", 
         "Средняя относительная ошибка", "Не рассчитывается, где y=0"),
        ("WAPE", "=IF(SUM(rngY)>0;SUM(rngAbsE)/SUM(rngY);NA())", "%", 
         "Взвешенная ошибка по объёму", "Устойчива к выбросам"),
        ("MAE", "=IFERROR(AVERAGEIF(rngAbsE;\">0\");NA())", "шт", 
         "Средняя абсолютная ошибка", "Простая интерпретация"),
        ("RMSE", "=IFERROR(SQRT(AVERAGE(rngE^2));NA())", "шт", 
         "Среднеквадратичная ошибка", "Усиливает влияние больших промахов"),
        ("Bias", "=IFERROR(AVERAGE(rngE);NA())", "шт", 
         "Систематическое смещение", "Знак = направление смещения")
    ]
    
    for i, (metric, formula, unit, interp, note) in enumerate(metrics_data, 4):
        ws_metrics.cell(row=i, column=1, value=metric).font = Font(bold=True)
        ws_metrics.cell(row=i, column=2, value=formula)
        ws_metrics.cell(row=i, column=3, value=unit)
        ws_metrics.cell(row=i, column=4, value=interp)
        ws_metrics.cell(row=i, column=5, value=note)
        
        # Форматирование
        ws_metrics.cell(row=i, column=2).fill = calc_fill
        if unit == "%":
            ws_metrics.cell(row=i, column=2).number_format = '0.00%'
        else:
            ws_metrics.cell(row=i, column=2).number_format = '0.00'
    
    # Именованные ячейки для метрик
    wb.defined_names['MAPE_value'] = DefinedName('MAPE_value', attr_text="Метрики!$B$4")
    wb.defined_names['WAPE_value'] = DefinedName('WAPE_value', attr_text="Метрики!$B$5")
    wb.defined_names['MAE_value'] = DefinedName('MAE_value', attr_text="Метрики!$B$6")
    wb.defined_names['RMSE_value'] = DefinedName('RMSE_value', attr_text="Метрики!$B$7")
    wb.defined_names['Bias_value'] = DefinedName('Bias_value', attr_text="Метрики!$B$8")
    
    # Диагностика
    ws_metrics['A10'] = 'ДИАГНОСТИКА'
    ws_metrics['A10'].font = Font(bold=True, size=12)
    ws_metrics.merge_cells('A10:E10')
    
    ws_metrics['A12'] = 'Знак Bias:'
    ws_metrics['B12'] = '=IF(Bias_value>0;"Недопрогноз (дефициты)";IF(Bias_value<0;"Перепрогноз (излишки)";"Смещения нет"))'
    ws_metrics['B12'].fill = calc_fill
    
    ws_metrics['A13'] = 'Сравнение RMSE vs MAE:'
    ws_metrics['B13'] = '=IF(RMSE_value>MAE_value;"Есть выбросы/крупные промахи";"Ошибки равномерны")'
    ws_metrics['B13'].fill = calc_fill
    
    # Рекомендации
    ws_metrics['A15'] = 'РЕКОМЕНДАЦИИ'
    ws_metrics['A15'].font = Font(bold=True, size=12)
    ws_metrics.merge_cells('A15:E15')
    
    ws_metrics['A17'] = 'При высоком WAPE:'
    ws_metrics['B17'] = '=IF(WAPE_value>0.15;"Проверьте сезонность/акции/качество данных";"")'
    ws_metrics['B17'].fill = calc_fill
    
    ws_metrics['A18'] = 'При RMSE >> MAE:'
    ws_metrics['B18'] = '=IF(RMSE_value>MAE_value*1.5;"Сгладьте выбросы или разнесите всплески";"")'
    ws_metrics['B18'].fill = calc_fill
    
    ws_metrics['A19'] = 'При значимом |Bias|:'
    ws_metrics['B19'] = '=IF(ABS(Bias_value)>MAE_value*0.3;"Сместите модель/перепараметризуйте";"")'
    ws_metrics['B19'].fill = calc_fill
    
    # Условное форматирование для WAPE
    ws_metrics.conditional_formatting.add('B5',
        CellIsRule(operator='greaterThan', formula=['0.15'], fill=PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type='solid')))
    ws_metrics.conditional_formatting.add('B5',
        CellIsRule(operator='between', formula=['0.10', '0.15'], fill=PatternFill(start_color=COLOR_AMBER, end_color=COLOR_AMBER, fill_type='solid')))
    ws_metrics.conditional_formatting.add('B5',
        CellIsRule(operator='lessThan', formula=['0.10'], fill=PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type='solid')))
    
    # Условное форматирование для Bias
    ws_metrics.conditional_formatting.add('B8',
        CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color=COLOR_BLUE, end_color=COLOR_BLUE, fill_type='solid')))
    ws_metrics.conditional_formatting.add('B8',
        CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill(start_color=COLOR_PURPLE, end_color=COLOR_PURPLE, fill_type='solid')))
    
    # Ширина колонок
    ws_metrics.column_dimensions['A'].width = 25
    ws_metrics.column_dimensions['B'].width = 20
    ws_metrics.column_dimensions['C'].width = 12
    ws_metrics.column_dimensions['D'].width = 30
    ws_metrics.column_dimensions['E'].width = 35
    
    # ============================================
    # Лист 4: Графики
    # ============================================
    ws_charts = wb.create_sheet("Графики", 3)
    
    ws_charts['A1'] = 'ВИЗУАЛИЗАЦИЯ МЕТРИК'
    ws_charts['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws_charts.merge_cells('A1:H1')
    ws_charts['A1'].alignment = center_align
    
    # График 1: Факт vs Прогноз (линейный)
    chart1 = LineChart()
    chart1.title = "Факт vs Прогноз"
    chart1.style = 12
    chart1.y_axis.title = 'Количество (шт)'
    chart1.x_axis.title = 'Период'
    chart1.height = 10
    chart1.width = 20
    
    # Данные для графика (первые 50 строк для примера)
    data1 = Reference(ws_data, min_col=3, min_row=1, max_row=50, max_col=4)
    cats1 = Reference(ws_data, min_col=2, min_row=2, max_row=50)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    
    ws_charts.add_chart(chart1, "A3")
    
    # График 2: Абсолютная ошибка (столбцы)
    chart2 = BarChart()
    chart2.type = 'col'
    chart2.title = "Абсолютная ошибка по периодам"
    chart2.style = 11
    chart2.y_axis.title = 'Абс. ошибка (шт)'
    chart2.x_axis.title = 'Период'
    chart2.height = 10
    chart2.width = 20
    
    data2 = Reference(ws_data, min_col=6, min_row=1, max_row=50)
    cats2 = Reference(ws_data, min_col=2, min_row=2, max_row=50)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    
    ws_charts.add_chart(chart2, "A20")
    
    # Карта статусов
    ws_charts['K3'] = 'КАРТА СТАТУСОВ'
    ws_charts['K3'].font = Font(bold=True, size=12)
    
    ws_charts['K5'] = 'WAPE:'
    ws_charts['L5'] = '=Метрики!B5'
    ws_charts['L5'].number_format = '0.00%'
    ws_charts['M5'] = '=IF(L5<0.10;"✓ Отлично";IF(L5<0.15;"⚠ Хорошо";"✗ Требуется улучшение"))'
    
    ws_charts['K6'] = 'Bias:'
    ws_charts['L6'] = '=Метрики!B8'
    ws_charts['L6'].number_format = '0.00'
    ws_charts['M6'] = '=IF(ABS(L6)<1;"✓ Без смещения";IF(L6>0;"⚠ Недопрогноз";"⚠ Перепрогноз"))'
    
    ws_charts['K7'] = 'RMSE/MAE:'
    ws_charts['L7'] = '=Метрики!B7/Метрики!B6'
    ws_charts['L7'].number_format = '0.00'
    ws_charts['M7'] = '=IF(L7>1.5;"⚠ Выбросы";"✓ Равномерно")'
    
    # ============================================
    # Тестовые данные (12 периодов)
    # ============================================
    test_data = [
        ("SKU-001", 1, 100, 92),
        ("SKU-001", 2, 105, 110),
        ("SKU-001", 3, 98, 95),
        ("SKU-001", 4, 110, 105),
        ("SKU-001", 5, 115, 120),
        ("SKU-001", 6, 120, 115),
        ("SKU-001", 7, 125, 130),
        ("SKU-001", 8, 118, 115),
        ("SKU-001", 9, 122, 125),
        ("SKU-001", 10, 130, 128),
        ("SKU-001", 11, 135, 140),
        ("SKU-001", 12, 140, 135),
    ]
    
    for i, (sku, period, fact, forecast) in enumerate(test_data, 2):
        ws_data.cell(row=i, column=1, value=sku)
        ws_data.cell(row=i, column=2, value=period)
        ws_data.cell(row=i, column=3, value=fact)
        ws_data.cell(row=i, column=4, value=forecast)
    
    # ============================================
    # Защита листов
    # ============================================
    # Защищаем лист "Метрики" (только просмотр)
    ws_metrics.protection.sheet = True
    ws_metrics.protection.password = ''
    ws_metrics.protection.enable()
    
    # Защищаем лист "Графики" (только просмотр)
    ws_charts.protection.sheet = True
    ws_charts.protection.password = ''
    ws_charts.protection.enable()
    
    # Защищаем лист "Ввод_данных" с разрешением редактирования C, D, H
    ws_data.protection.sheet = True
    ws_data.protection.password = ''
    
    # Разблокируем ячейки для ввода
    for row in range(2, 1001):
        ws_data.cell(row=row, column=1).protection = openpyxl.styles.Protection(locked=False)  # SKU
        ws_data.cell(row=row, column=2).protection = openpyxl.styles.Protection(locked=False)  # Период
        ws_data.cell(row=row, column=3).protection = openpyxl.styles.Protection(locked=False)  # Факт
        ws_data.cell(row=row, column=4).protection = openpyxl.styles.Protection(locked=False)  # Прогноз
        ws_data.cell(row=row, column=8).protection = openpyxl.styles.Protection(locked=False)  # Примечание
    
    ws_data.protection.enable()
    
    # Сохраняем файл
    output_file = '/workspace/Forecast_Metrics_Template.xlsx'
    wb.save(output_file)
    print(f"✅ Файл успешно создан: {output_file}")
    
    return output_file

if __name__ == "__main__":
    create_forecast_template()
